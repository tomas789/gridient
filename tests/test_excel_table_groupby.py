# tests/test_excel_table_groupby.py

import os
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import gridient as gr


# Assuming the same helper function as in test_excel_series_slicing.py
# Or reuse/import from a central test utility module
def write_and_read_workbook(layout: gr.ExcelLayout, filename="test_output.xlsx"):
    """Writes the layout to an Excel file and returns the loaded openpyxl workbook."""
    output_path = Path(filename)
    if output_path.exists():
        os.remove(output_path)

    layout.write()

    if not output_path.exists():
        raise FileNotFoundError(f"Test workbook '{output_path}' was not created.")

    try:
        workbook = openpyxl.load_workbook(str(output_path), data_only=False)
    except Exception as e:
        raise IOError(f"Failed to read back workbook '{output_path}': {e}")
    finally:
        if output_path.exists():
            os.remove(output_path)

    return workbook


# --- Groupby / Pivot Table Tests ---


def test_table_groupby_sum_single_column():
    """Test table.groupby('Category').sum()"""
    filename = "test_groupby_sum.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    # Original Data
    data = {"Category": ["A", "B", "A", "B", "A"], "Value": [10, 20, 15, 25, 12]}
    df = pd.DataFrame(data)
    # Create ExcelSeries from pandas Series
    category_series = gr.ExcelSeries.from_pandas(df["Category"], name="Category")
    value_series = gr.ExcelSeries.from_pandas(df["Value"], name="Value")
    # Correct way to initialize ExcelTable
    original_table = gr.ExcelTable(columns=[category_series, value_series], title="SourceData")

    # Place original table (header in row 1, data from row 2)
    # Let's assume header=True is default or handled by from_pandas
    # Placing at A1 implies header in A1:B1, data in A2:B6
    sheet1.add(original_table, row=1, col=1)  # A1

    # Perform groupby and sum
    # This assumes .groupby().sum() returns a new ExcelTable
    pivot_table = original_table.groupby("Category").sum()
    pivot_table.name = "PivotSum"

    # Place pivot table (e.g., starting D1 for header)
    # Header D1:E1, Index A in D2, B in D3
    # Values in E2, E3
    sheet1.add(pivot_table, row=1, col=4)  # D1

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Assertions
    # Check Pivot Headers
    assert sheet_xl["D1"].value == "Category"
    assert sheet_xl["E1"].value == "Value"  # Assuming 'Value' is the sum column name

    # Check Pivot Index
    assert sheet_xl["D2"].value == "A"
    assert sheet_xl["D3"].value == "B"

    # Check Pivot Value Formulas (Most critical part)
    # Formula should reference the original table's data range (A2:B6)
    # It needs the category column (A2:A6) and value column (B2:B6)
    # The criteria comes from the pivot table's index column (D2 for 'A', D3 for 'B')
    # Note: Using $ for absolute row references in ranges is good practice for formulas
    #       that might be conceptually copied down.
    assert sheet_xl["E2"].value == "=SUMIF(A$2:A$6,D2,B$2:B$6)"  # Sum for 'A'
    assert sheet_xl["E3"].value == "=SUMIF(A$2:A$6,D3,B$2:B$6)"  # Sum for 'B'

    # Optional: Check original data
    assert sheet_xl["A1"].value == "Category"
    assert sheet_xl["B1"].value == "Value"
    assert sheet_xl["A2"].value == "A"
    assert sheet_xl["B2"].value == 10


def test_table_groupby_mean_single_column():
    """Test table.groupby('Category').mean()"""
    filename = "test_groupby_mean.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = {"Category": ["A", "B", "A", "B", "A"], "Value": [10, 20, 15, 25, 12]}
    df = pd.DataFrame(data)
    # Create ExcelSeries from pandas Series
    category_series = gr.ExcelSeries.from_pandas(df["Category"], name="Category")
    value_series = gr.ExcelSeries.from_pandas(df["Value"], name="Value")
    # Correct way to initialize ExcelTable
    original_table = gr.ExcelTable(columns=[category_series, value_series], title="SourceData")
    sheet1.add(original_table, row=1, col=1)  # A1 (Data A2:B6)

    pivot_table = original_table.groupby("Category").mean()
    pivot_table.name = "PivotMean"
    sheet1.add(pivot_table, row=1, col=4)  # D1 (Index D2:D3, Values E2:E3)

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Check Pivot Value Formulas
    assert sheet_xl["E2"].value == "=AVERAGEIF(A$2:A$6,D2,B$2:B$6)"  # Mean for 'A'
    assert sheet_xl["E3"].value == "=AVERAGEIF(A$2:A$6,D3,B$2:B$6)"  # Mean for 'B'


def test_table_groupby_count_single_column():
    """Test table.groupby('Category').count()"""
    filename = "test_groupby_count.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    # Note: Count might apply to the category column itself or value, TBD
    # Let's assume it counts occurrences of the category
    data = {"Category": ["A", "B", "A", "B", "A"], "Value": [10, 20, 15, 25, 12]}
    df = pd.DataFrame(data)
    # Create ExcelSeries from pandas Series
    category_series = gr.ExcelSeries.from_pandas(df["Category"], name="Category")
    value_series = gr.ExcelSeries.from_pandas(df["Value"], name="Value")
    # Correct way to initialize ExcelTable
    original_table = gr.ExcelTable(columns=[category_series, value_series], title="SourceData")
    sheet1.add(original_table, row=1, col=1)  # A1 (Data A2:B6)

    # Assuming count counts the non-null values per group in one of the columns.
    # Let's assume it returns counts for all original columns (excluding group key)
    # or maybe just a single count column. Let's assume it counts 'Category' occurrences.
    pivot_table = original_table.groupby("Category").count()  # Behavior needs definition
    pivot_table.name = "PivotCount"
    # If it counts 'Category' occurrences, pivot might have columns 'Category' and 'count'
    sheet1.add(pivot_table, row=1, col=4)  # D1 (Index D2:D3, Values E2:E3)

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Assertions depend heavily on how `count()` is implemented.
    # Scenario 1: Counts occurrences of the group key in the group key column
    assert sheet_xl["E1"].value == "count"  # Or maybe 'Category_count' or just 'Value' if it counts that
    assert sheet_xl["E2"].value == "=COUNTIF(A$2:A$6,D2)"  # Count for 'A'
    assert sheet_xl["E3"].value == "=COUNTIF(A$2:A$6,D3)"  # Count for 'B'


@pytest.mark.skip(reason="Groupby multiple columns TBD")
def test_table_groupby_sum_multiple_columns():
    """Test table.groupby(['Cat1', 'Cat2']).sum()"""
    filename = "test_groupby_sum_multi.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = {"Cat1": ["A", "A", "B", "B", "A"], "Cat2": ["X", "Y", "X", "Y", "X"], "Value": [10, 20, 15, 25, 12]}
    df = pd.DataFrame(data)
    # Create ExcelSeries from pandas Series
    cat1_series = gr.ExcelSeries.from_pandas(df["Cat1"], name="Cat1")
    cat2_series = gr.ExcelSeries.from_pandas(df["Cat2"], name="Cat2")
    value_series = gr.ExcelSeries.from_pandas(df["Value"], name="Value")
    # Correct way to initialize ExcelTable
    original_table = gr.ExcelTable(columns=[cat1_series, cat2_series, value_series], title="SourceDataMulti")
    # Place original table (Header A1:C1, Data A2:C6)
    sheet1.add(original_table, row=1, col=1)  # A1

    pivot_table = original_table.groupby(["Cat1", "Cat2"]).sum()
    pivot_table.name = "PivotSumMulti"
    # Place pivot table (e.g., starting E1)
    # Header E1:G1 (Cat1, Cat2, Value), Data E2:G?
    sheet1.add(pivot_table, row=1, col=5)  # E1

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Assertions
    # Check Headers
    assert sheet_xl["E1"].value == "Cat1"
    assert sheet_xl["F1"].value == "Cat2"
    assert sheet_xl["G1"].value == "Value"

    # Check Index Values (e.g., find the row for A, X)
    # This requires finding the row - let's assume A,X is in row 2 (E2="A", F2="X")
    # Find the row corresponding to group ('A', 'X') in the pivot table output
    pivot_row = -1
    for row_idx in range(2, sheet_xl.max_row + 1):
        if sheet_xl[f"E{row_idx}"].value == "A" and sheet_xl[f"F{row_idx}"].value == "X":
            pivot_row = row_idx
            break
    assert pivot_row != -1, "Could not find pivot row for ('A', 'X')"

    # Check Pivot Value Formula using SUMIFS
    # Sum 'Value' (C2:C6) where 'Cat1' (A2:A6) matches E<row> AND 'Cat2' (B2:B6) matches F<row>
    expected_formula = f"=SUMIFS(C$2:C$6, A$2:A$6, E{pivot_row}, B$2:B$6, F{pivot_row})"
    assert sheet_xl[f"G{pivot_row}"].value == expected_formula
