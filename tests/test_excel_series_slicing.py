# tests/test_excel_series_slicing.py

import os
from pathlib import Path

import openpyxl  # Assuming openpyxl is used for reading back values/formulas
import pytest

import gridient as gr


# Helper function (potentially adapt from other test files or create anew)
def write_and_read_workbook(layout: gr.ExcelLayout, filename="test_output.xlsx"):
    """Writes the layout to an Excel file and returns the loaded openpyxl workbook."""
    # Ensure the output directory exists if needed, or manage paths
    output_path = Path(filename)
    if output_path.exists():
        os.remove(output_path)

    layout.write()

    # Check if file was created
    if not output_path.exists():
        raise FileNotFoundError(f"Test workbook '{output_path}' was not created.")

    # Read back using openpyxl
    try:
        workbook = openpyxl.load_workbook(str(output_path), data_only=False)  # data_only=False to read formulas
    except Exception as e:
        raise IOError(f"Failed to read back workbook '{output_path}': {e}")
    finally:
        # Clean up the file
        if output_path.exists():
            os.remove(output_path)

    return workbook


# --- Slicing Tests ---


def test_series_basic_slice_sum():
    """Test summing a basic slice like series[1:4]."""
    filename = "test_slice_sum.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = [10, 20, 30, 40, 50]
    series = gr.ExcelSeries(data=data, name="Data")

    # Add series starting at B2 (Excel index)
    sheet1.add(series, row=2, col=2)  # B2

    # Create slice and sum it
    series_slice = series[1:4]  # Should represent [20, 30, 40] -> cells B3:B5
    total = gr.sum(series_slice)  # Use gridient's sum
    total.name = "Slice Sum"

    # Add the sum result at D2
    sheet1.add(total, row=2, col=4)  # D2

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Assertion: Check the formula in D2
    assert sheet_xl["D2"].value == "=SUM(B3:B5)"
    # Optional: Check source data was written correctly
    assert sheet_xl["B2"].value == 10
    assert sheet_xl["B3"].value == 20
    assert sheet_xl["B4"].value == 30
    assert sheet_xl["B5"].value == 40
    assert sheet_xl["B6"].value == 50


def test_series_slice_from_start_sum():
    """Test summing a slice from the start like series[:2]."""
    filename = "test_slice_start_sum.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = [10, 20, 30, 40, 50]
    series = gr.ExcelSeries(data=data, name="Data")
    sheet1.add(series, row=2, col=2)  # B2

    series_slice = series[:2]  # Should represent [10, 20] -> cells B2:B3
    total = gr.sum(series_slice)
    total.name = "Slice Sum Start"
    sheet1.add(total, row=2, col=4)  # D2

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    assert sheet_xl["D2"].value == "=SUM(B2:B3)"


def test_series_slice_to_end_sum():
    """Test summing a slice to the end like series[3:]."""
    filename = "test_slice_end_sum.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = [10, 20, 30, 40, 50]
    series = gr.ExcelSeries(data=data, name="Data")
    sheet1.add(series, row=2, col=2)  # B2

    series_slice = series[3:]  # Should represent [40, 50] -> cells B5:B6
    total = gr.sum(series_slice)
    total.name = "Slice Sum End"
    sheet1.add(total, row=2, col=4)  # D2

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    assert sheet_xl["D2"].value == "=SUM(B5:B6)"


@pytest.mark.skip(reason="Single element slicing behaviour TBD")
def test_series_single_element_slice_sum():
    """Test summing a slice representing a single element series[2:3]."""
    filename = "test_single_slice_sum.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    data = [10, 20, 30, 40, 50]
    series = gr.ExcelSeries(data=data, name="Data")
    sheet1.add(series, row=2, col=2)  # B2

    series_slice = series[2:3]  # Should represent [30] -> cell B4
    total = gr.sum(series_slice)  # Could result in =SUM(B4) or just =B4
    total.name = "Single Slice Sum"
    sheet1.add(total, row=2, col=4)  # D2

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Decide expected behavior: SUM function or direct reference?
    # assert sheet_xl['D2'].value == "=SUM(B4)"
    assert sheet_xl["D2"].value == "=B4"  # Assuming direct reference for single cell


@pytest.mark.skip(reason="Element-wise slice operations TBD")
def test_series_slice_elementwise_addition():
    """Test adding a slice to another series element-wise."""
    filename = "test_slice_elementwise_add.xlsx"
    workbook_gr = gr.ExcelWorkbook(filename)
    layout = gr.ExcelLayout(workbook_gr)
    sheet1 = gr.ExcelSheetLayout("Sheet1")

    s1_data = [10, 20, 30, 40, 50]
    s2_data = [1, 2, 3]  # Shorter series for addition

    s1 = gr.ExcelSeries(data=s1_data, name="S1")
    s2 = gr.ExcelSeries(data=s2_data, name="S2")

    sheet1.add(s1, row=2, col=2)  # B2:B6
    sheet1.add(s2, row=2, col=4)  # D2:D4

    s1_slice = s1[1:4]  # Represents [20, 30, 40] -> cells B3:B5

    # Add the slice (B3:B5) to s2 (D2:D4)
    result_series = s1_slice + s2
    result_series.name = "Slice Add Result"

    sheet1.add(result_series, row=2, col=6)  # F2:F4

    layout.add_sheet(sheet1)
    workbook_xl = write_and_read_workbook(layout, filename)
    sheet_xl = workbook_xl["Sheet1"]

    # Assertions: Check formulas in F2, F3, F4
    assert sheet_xl["F2"].value == "=B3+D2"  # 20 + 1
    assert sheet_xl["F3"].value == "=B4+D3"  # 30 + 2
    assert sheet_xl["F4"].value == "=B5+D4"  # 40 + 3
