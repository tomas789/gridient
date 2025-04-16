import openpyxl

from gridient import ExcelLayout, ExcelSheetLayout, ExcelValue, ExcelWorkbook
from gridient.tables import ExcelParameterTable
from gridient.values import ExcelSeries


def check_formula_in_xlsx(filename, sheet_name, row, col):
    """Helper to extract formula from an Excel cell using openpyxl."""
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    cell = ws.cell(row=row + 1, column=col + 1)  # openpyxl uses 1-based indexing
    return cell.value if cell.value and cell.value.startswith("=") else None


def test_cross_sheet_value_reference():
    """Test if a value reference from one sheet to another includes the sheet name."""
    # Create a workbook with two sheets
    filename = "test_cross_sheet_value.xlsx"
    workbook = ExcelWorkbook(filename)
    layout = ExcelLayout(workbook)

    # Create two sheets
    sheet1 = ExcelSheetLayout("Sheet1")
    sheet2 = ExcelSheetLayout("Sheet2")

    # Add value to first sheet
    value1 = ExcelValue(42, name="Test Value")
    sheet1.add(value1, 1, 1)

    # Create formula referencing the value in second sheet
    formula = ExcelValue(value1 * 2)  # 2 * value1
    sheet2.add(formula, 2, 2)

    # Add sheets to layout
    layout.add_sheet(sheet1)
    layout.add_sheet(sheet2)

    # Write workbook
    workbook.layout = layout
    layout.write()

    # Check if the formula in sheet2 includes a reference to Sheet1
    formula_text = check_formula_in_xlsx(filename, "Sheet2", 2, 2)
    assert formula_text is not None

    # Formula should contain Sheet1 reference
    assert "Sheet1" in formula_text, f"Sheet name not found in formula: {formula_text}"


def test_cross_sheet_parameter_reference():
    """Test if a parameter reference from one sheet to another includes the sheet name."""
    # Create a workbook with two sheets
    filename = "test_cross_sheet_parameter.xlsx"
    workbook = ExcelWorkbook(filename)
    layout = ExcelLayout(workbook)

    # Create two sheets
    sheet1 = ExcelSheetLayout("Sheet1")
    sheet2 = ExcelSheetLayout("Sheet2")

    # Add parameter to first sheet (using is_parameter=True flag)
    param1 = ExcelValue(100, name="Test Param", is_parameter=True)
    sheet1.add(param1, 1, 1)

    # Create formula referencing the parameter in second sheet
    formula = ExcelValue(param1 * 2)  # 2 * param1
    sheet2.add(formula, 2, 2)

    # Add sheets to layout
    layout.add_sheet(sheet1)
    layout.add_sheet(sheet2)

    # Write workbook
    workbook.layout = layout
    layout.write()

    # Check if the formula in sheet2 includes a reference to Sheet1
    formula_text = check_formula_in_xlsx(filename, "Sheet2", 2, 2)
    assert formula_text is not None

    # Formula should contain Sheet1 reference
    assert "Sheet1" in formula_text, f"Sheet name not found in formula: {formula_text}"


def test_cross_sheet_parameter_table_reference():
    """Test if a parameter table reference from one sheet to another includes the sheet name."""
    # Create a workbook with two sheets
    filename = "test_cross_sheet_table.xlsx"
    workbook = ExcelWorkbook(filename)
    layout = ExcelLayout(workbook)

    # Create two sheets
    sheet1 = ExcelSheetLayout("Sheet1")
    sheet2 = ExcelSheetLayout("Sheet2")

    # Create parameter table in first sheet
    table = ExcelParameterTable()
    param1 = ExcelValue(10, name="param1", unit="Unit1", is_parameter=True)
    param2 = ExcelValue(20, name="param2", unit="Unit2", is_parameter=True)
    table.add(param1)
    table.add(param2)

    sheet1.add(table, 1, 1)

    # Create formula referencing the parameter in second sheet
    formula = ExcelValue(param1 * 3)  # 3 * param1
    sheet2.add(formula, 2, 2)

    # Add sheets to layout
    layout.add_sheet(sheet1)
    layout.add_sheet(sheet2)

    # Write workbook
    workbook.layout = layout
    layout.write()

    # Check if the formula in sheet2 includes a reference to Sheet1
    formula_text = check_formula_in_xlsx(filename, "Sheet2", 2, 2)
    assert formula_text is not None

    # Formula should contain Sheet1 reference
    assert "Sheet1" in formula_text, f"Sheet name not found in formula: {formula_text}"


def test_cross_sheet_series_reference():
    """Test if a series reference from one sheet to another includes the sheet name."""
    # Create a workbook with two sheets
    filename = "test_cross_sheet_series.xlsx"
    workbook = ExcelWorkbook(filename)
    layout = ExcelLayout(workbook)

    # Create two sheets
    sheet1 = ExcelSheetLayout("Sheet1")
    sheet2 = ExcelSheetLayout("Sheet2")

    # Create series in first sheet
    series = ExcelSeries(name="Test Series")
    series[0] = 100
    series[1] = 200
    series[2] = 300

    sheet1.add(series, 1, 1)

    # Access element and create formula in second sheet
    element = series[1]  # Value 200
    formula = ExcelValue(element * 2)  # 2 * 200
    sheet2.add(formula, 2, 2)

    # Add sheets to layout
    layout.add_sheet(sheet1)
    layout.add_sheet(sheet2)

    # Write workbook
    workbook.layout = layout
    layout.write()

    # Check if the formula in sheet2 includes a reference to Sheet1
    formula_text = check_formula_in_xlsx(filename, "Sheet2", 2, 2)
    assert formula_text is not None

    # Formula should contain Sheet1 reference
    assert "Sheet1" in formula_text, f"Sheet name not found in formula: {formula_text}"
